"""Excel report generator using openpyxl."""
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Marsh brand colors
COBALT = "002C77"
TEAL = "00A8C7"
EMERALD = "2DD4A8"
CORAL = "F97066"
WHITE = "FFFFFF"
LIGHT_GRAY = "F4F7FC"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=COBALT, end_color=COBALT, fill_type="solid")
TEAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
RED_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CURRENCY_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D4DCE8"),
)


class XLSXGenerator:
    """Generates formatted Excel reports with Marsh branding."""

    async def generate(self, context: Any) -> bytes:
        """Generate XLSX from ReportContext. Returns bytes."""
        wb = Workbook()

        # Sheet 1: Summary
        self._build_summary_sheet(wb.active, context)

        # Sheet 2: All Variances
        ws_var = wb.create_sheet("Material Variances")
        self._build_variance_sheet(ws_var, context.variances)

        # Per-BU tabs
        bu_ids = sorted(set(v.get("bu_id", "") for v in context.variances if v.get("bu_id")))
        for bu_id in bu_ids[:5]:
            bu_vars = [v for v in context.variances if v.get("bu_id") == bu_id]
            if bu_vars:
                name = bu_id.replace("_", " ").title()[:31]
                ws_bu = wb.create_sheet(name)
                ws_bu.sheet_properties.tabColor = TEAL
                self._build_variance_sheet(ws_bu, bu_vars)

        # P&L sheet
        if context.pl_rows:
            ws_pl = wb.create_sheet("P&L")
            self._build_pl_sheet(ws_pl, context.pl_rows)

        # Write to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_summary_sheet(self, ws, context) -> None:
        ws.title = "Summary"
        ws.sheet_properties.tabColor = COBALT

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Variance Analysis — {context.period_id} {context.view} vs {context.comparison_base}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=COBALT)
        ws["A1"].alignment = Alignment(horizontal="center")

        # KPI Table
        headers = ["Metric", "Actual", "Comparator", "Variance $", "Variance %", "Favorable"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for i, card in enumerate(context.summary_cards):
            row = 4 + i
            ws.cell(row=row, column=1, value=card.get("metric_name", ""))
            ws.cell(row=row, column=2, value=card.get("actual", 0)).number_format = CURRENCY_FMT
            ws.cell(row=row, column=3, value=card.get("comparator", 0)).number_format = CURRENCY_FMT
            var_cell = ws.cell(row=row, column=4, value=card.get("variance_amount", 0))
            var_cell.number_format = CURRENCY_FMT
            pct = card.get("variance_pct", 0)
            ws.cell(row=row, column=5, value=pct / 100 if pct else 0).number_format = PCT_FMT
            fav = card.get("is_favorable", False)
            ws.cell(row=row, column=6, value="Y" if fav else "N")
            # Conditional fill
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = GREEN_FILL if fav else RED_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER

        # Auto-width
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _build_variance_sheet(self, ws, variances: list[dict]) -> None:
        headers = ["Account", "BU", "Geo", "Variance $", "Variance %", "Type", "Status", "Narrative"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Sort by absolute variance
        sorted_vars = sorted(variances, key=lambda v: abs(v.get("variance_amount", 0)), reverse=True)

        for i, v in enumerate(sorted_vars):
            row = 2 + i
            ws.cell(row=row, column=1, value=v.get("account_name", v.get("account_id", "")))
            ws.cell(row=row, column=2, value=v.get("bu_id", "").replace("_", " ").title())
            ws.cell(row=row, column=3, value=v.get("geo_node_id", ""))
            ws.cell(row=row, column=4, value=v.get("variance_amount", 0)).number_format = CURRENCY_FMT
            pct = v.get("variance_pct", 0)
            ws.cell(row=row, column=5, value=pct / 100 if pct else 0).number_format = PCT_FMT
            ws.cell(row=row, column=6, value=v.get("is_material") and "Material" or v.get("is_trending") and "Trending" or "")
            ws.cell(row=row, column=7, value=v.get("narrative_source", ""))
            ws.cell(row=row, column=8, value=v.get("narrative_oneliner", ""))

            fav = v.get("variance_amount", 0) > 0  # Simplified
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = THIN_BORDER
                if abs(v.get("variance_pct", 0) or 0) > 5:
                    ws.cell(row=row, column=col).fill = GREEN_FILL if fav else RED_FILL

        # Freeze panes + auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{len(sorted_vars) + 1}"

        # Column widths
        widths = [30, 18, 15, 15, 12, 12, 12, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _build_pl_sheet(self, ws, pl_rows: list[dict]) -> None:
        headers = ["Account", "Actual", "Comparator", "Variance $", "Variance %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        def flatten(rows, depth=0, output=None):
            if output is None:
                output = []
            for r in rows if isinstance(rows, list) else [rows]:
                output.append((r, depth))
                for child in r.get("children", []):
                    flatten([child], depth + 1, output)
            return output

        flat = flatten(pl_rows)
        for i, (r, depth) in enumerate(flat):
            row = 2 + i
            name = r.get("account_name", "")
            is_calc = r.get("is_calculated", False)
            ws.cell(row=row, column=1, value=("  " * depth) + name).font = BOLD_FONT if is_calc else NORMAL_FONT
            ws.cell(row=row, column=2, value=r.get("actual", 0)).number_format = CURRENCY_FMT
            ws.cell(row=row, column=3, value=r.get("comparator", 0)).number_format = CURRENCY_FMT
            ws.cell(row=row, column=4, value=r.get("variance_amount", 0)).number_format = CURRENCY_FMT
            pct = r.get("variance_pct", 0)
            ws.cell(row=row, column=5, value=pct / 100 if pct else 0).number_format = PCT_FMT

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25 if col == 1 else 18
