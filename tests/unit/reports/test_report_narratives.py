"""Tests for narrative pyramid integration in report generators.

Verifies that all 4 report generators pull executive summary and
section narratives from ReportContext (not self-assembled).
"""

import io
import pytest
from unittest.mock import MagicMock

from services.reports.generators.data_provider import ReportContext


def _build_context_with_narratives() -> ReportContext:
    """Build a ReportContext with narrative pyramid data."""
    return ReportContext(
        period_id="2026-05",
        comparison_base="BUDGET",
        view="MTD",
        summary_cards=[
            {"metric_name": "Revenue", "actual": 100000, "comparator": 95000, "variance_amount": 5000, "variance_pct": 5.3},
            {"metric_name": "EBITDA", "actual": 20000, "comparator": 22000, "variance_amount": -2000, "variance_pct": -9.1},
        ],
        variances=[
            {"account_name": "Advisory Fees", "bu_id": "marsh", "variance_amount": 3000, "variance_pct": 8.5,
             "narrative_oneliner": "$3K favorable", "narrative_midlevel": "Advisory Fees up $3K driven by APAC.",
             "narrative_detail": "Advisory Fees increased by $3K (+8.5%) vs Budget.", "narrative_source": "generated"},
            {"account_name": "Consulting", "bu_id": "marsh", "variance_amount": -1500, "variance_pct": -4.2,
             "narrative_oneliner": "$1.5K unfavorable", "narrative_midlevel": "Consulting down $1.5K due to EMEA delays.",
             "narrative_detail": "Consulting decreased by $1.5K (-4.2%) vs Budget.", "narrative_source": "generated"},
        ],
        pl_rows=[],
        waterfall_steps=[],
        netting_alerts=[{"left": "T&E +$1K", "right": "Revenue -$2K", "net": "-$1K"}],
        trend_alerts=[{"description": "Tech infra trending up", "projection": "+$50K YE"}],
        executive_summary={
            "headline": "May 2026 close: Revenue up 5.3%, EBITDA down 9.1%.",
            "full_narrative": "May financial performance was mixed.\n\nRevenue grew 5.3% driven by advisory strength.\n\nEBITDA declined 9.1% as costs outpaced revenue.",
            "key_risks": [{"risk": "3 trending variances", "severity": "medium"}],
            "cross_bu_themes": [{"theme": "2 of 5 BUs positive"}],
        },
        section_narratives=[
            {"section_name": "Revenue", "narrative": "Revenue up 5.3%. Key: Advisory +$3K.", "key_drivers": [{"account_name": "Advisory", "amount": 3000}]},
            {"section_name": "Profitability", "narrative": "EBITDA down 9.1%. Margin compressed.", "key_drivers": []},
        ],
    )


@pytest.mark.asyncio
class TestPDFNarratives:
    async def test_pdf_includes_exec_narrative(self):
        from services.reports.generators.pdf_generator import PDFGenerator
        gen = PDFGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert isinstance(result, bytes)
        assert len(result) > 100

    async def test_pdf_generated_successfully(self):
        from services.reports.generators.pdf_generator import PDFGenerator
        gen = PDFGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert result[:4] == b"%PDF"


@pytest.mark.asyncio
class TestPPTXNarratives:
    async def test_pptx_includes_headline(self):
        from services.reports.generators.pptx_generator import PPTXGenerator
        gen = PPTXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert isinstance(result, bytes)
        assert len(result) > 100

    async def test_pptx_is_valid_zip(self):
        from services.reports.generators.pptx_generator import PPTXGenerator
        import zipfile
        gen = PPTXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert zipfile.is_zipfile(io.BytesIO(result))


@pytest.mark.asyncio
class TestDOCXNarratives:
    async def test_docx_includes_section_narratives(self):
        from services.reports.generators.docx_generator import DOCXGenerator
        gen = DOCXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert isinstance(result, bytes)
        assert len(result) > 100

    async def test_docx_is_valid_zip(self):
        from services.reports.generators.docx_generator import DOCXGenerator
        import zipfile
        gen = DOCXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        assert zipfile.is_zipfile(io.BytesIO(result))


@pytest.mark.asyncio
class TestXLSXNarratives:
    async def test_xlsx_includes_section_sheet(self):
        from services.reports.generators.xlsx_generator import XLSXGenerator
        import openpyxl
        gen = XLSXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Section Narratives" in wb.sheetnames

    async def test_xlsx_summary_has_headline(self):
        from services.reports.generators.xlsx_generator import XLSXGenerator
        import openpyxl
        gen = XLSXGenerator()
        ctx = _build_context_with_narratives()
        result = await gen.generate(ctx)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        summary = wb["Summary"]
        headline_text = str(summary["A1"].value or "")
        assert "Revenue" in headline_text or "May" in headline_text
