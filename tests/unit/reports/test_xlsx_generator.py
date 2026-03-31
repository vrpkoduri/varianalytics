"""Tests for XLSX report generator."""
import pytest
from services.reports.generators.xlsx_generator import XLSXGenerator
from services.reports.generators.data_provider import ReportContext


@pytest.fixture
def sample_context():
    return ReportContext(
        period_id="2026-06",
        summary_cards=[
            {"metric_name": "Revenue", "actual": 938000, "comparator": 951000, "variance_amount": -13000, "variance_pct": -1.38, "is_favorable": False},
            {"metric_name": "EBITDA", "actual": 158000, "comparator": 175000, "variance_amount": -17000, "variance_pct": -9.88, "is_favorable": False},
        ],
        variances=[
            {"account_id": "acct_advisory", "account_name": "Advisory Fees", "bu_id": "marsh", "geo_node_id": "geo_apac", "variance_amount": 6900, "variance_pct": 15.3, "is_material": True, "narrative_source": "llm", "narrative_oneliner": "$6.9K increased"},
            {"account_id": "acct_tech", "account_name": "Tech Infrastructure", "bu_id": "marsh", "geo_node_id": "geo_global", "variance_amount": -5800, "variance_pct": -8.3, "is_trending": True, "narrative_source": "llm", "narrative_oneliner": "$5.8K decreased"},
        ],
        pl_rows=[{"account_name": "Revenue", "actual": 938000, "comparator": 951000, "variance_amount": -13000, "is_calculated": False, "children": []}],
    )


class TestXLSXGenerator:
    @pytest.mark.asyncio
    async def test_generates_valid_xlsx(self, sample_context):
        gen = XLSXGenerator()
        data = await gen.generate(sample_context)
        assert isinstance(data, bytes)
        assert len(data) > 1000  # Non-trivial size
        # XLSX magic bytes: PK (ZIP format)
        assert data[:2] == b'PK'

    @pytest.mark.asyncio
    async def test_has_summary_sheet(self, sample_context):
        from openpyxl import load_workbook
        import io
        gen = XLSXGenerator()
        data = await gen.generate(sample_context)
        wb = load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_has_variance_sheet(self, sample_context):
        from openpyxl import load_workbook
        import io
        gen = XLSXGenerator()
        data = await gen.generate(sample_context)
        wb = load_workbook(io.BytesIO(data))
        assert "Material Variances" in wb.sheetnames
